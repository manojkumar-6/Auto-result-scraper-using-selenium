from .models import *
from django.conf import settings
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from re import S
import selenium
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as vc
from attr import s
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.http import HttpResponse
from importlib.resources import path
from pathlib import Path
from h11 import Data
from numpy import gradient
from pyparsing import col
import selenium
import time
import pandas as pd
from selenium import webdriver
import openpyxl
from pandas import DataFrame
from collections import defaultdict
from openpyxl import Workbook,load_workbook
from selenium.common import exceptions
from openpyxl.chart import BarChart,Reference
def h(request):
    f=mg.objects.all()
    return render(request,"a.html",{"f":f})
def r(request):
    ce=[]
    we=request.POST["u"]
    op=request.POST["s"]
    ui=request.POST["d"]
    hj=request.POST["ht"]
    sp=request.POST["sp"]
    cvb=request.POST["cvb"]
    ert=request.POST["ert"]
    ert=ert.split(",")
    ery=[]
    lkj=[]
    gkn="RMR501712.xlsx"
    for i in ert:
        ery.append(float(i))
    bo=set()
    we=str(we)
    op=int(op)
    if "," in ui:
        ui=ui.split(",")
        for i in ui:
            if i==",":
                continue
            else:
                ce.append(int(i))
    else:
        ce.append(int(ui))
    hj=str(hj)
    jk={"A":9,"S":10,"B":8,"C":7,"D":6,"E":5,"F":0,"Y":0}
    fg=dict()
    for i in range(1,len(ery)+1):
        fg[i]=ery[i-1]
    tr=dict()
    ak=[]
    k=[]
    zx={"P":1,"F":0,"AB":0}
    v=dict()
    l=list()
    wb=Workbook()
    ws=wb.active
    for i in range(1,16):
        tr[i]=0
    q=op
    sdf=[]
    for i in range(q,128):
        if i < 10 and i not in ce:
            j=hj+"0"+str(i)
            k.append(j)
        elif i in ce:
            continue
        elif i>=100 and i<110:
            j=hj+"A"+str(i-100)
            k.append(j)
        elif i>=110 and i<120:
            j=hj+"B"+str(i-110)
            k.append(j)
        elif i>=120 and i<130:
            j=hj+"C"+str(i-120)
            k.append(j)
        elif i>=130 and i<140:
            j=hj+"D"+str(i-130)
            k.append(j)
        else:
            j=hj+str(i)
            k.append(j)
    PATH = "C:\Program Files (x86)\chromedriver.exe"
    d=webdriver.Chrome(PATH)
    d.get(we)
    q1=wb.create_sheet("sheet1")
    q2=wb.create_sheet("sheet2")
    q3=wb.create_sheet("sheet3")
    q4=wb.create_sheet("sheet4")
    q5=wb.create_sheet("sheet5")
    q6=wb.create_sheet("sheet6")
    q7=wb.create_sheet("sheet7")
    q8=wb.create_sheet("sheet8")
    q9=wb.create_sheet("sheet9")
    q10=wb.create_sheet("sheet10")
    q11=wb.create_sheet("sheet11")
    q12=wb.create_sheet("sheet12")
    q13=wb.create_sheet("sheet13")
    q14=wb.create_sheet("sheet14")
    q15=wb.create_sheet("sheet15")
    q16=wb.create_sheet("sheet16")
    q17=wb.create_sheet("ranking")
    q1.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q2.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q3.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q4.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q5.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q6.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q7.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q8.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q9.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q10.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q11.append(['subject',"total number of failures","total pass strength"])
    q12.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q13.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q14.append(["hallticketnumber",'subject','internals','externals','total','resultstatus',"grade","credits"])
    q15.append(["hallticketnumber",'gpa'])
    q16.append(["subject","hallticketnumber"])
    q17.append(["rank","gpa","hallticketnumber"])
    zvbm=0
    hjk=[]
    dert=[]
    for i in k:
        opi=0
        es=0
        zf=[]
        jka=[]
        jka.append(i)
        t=d.find_element_by_id("ht")
        time.sleep(1.2)
        t.send_keys(i)
       # e=WebDriverWait(d,2).until(vc.presence_of_element_located(By.CLASS_NAME,"ci"))
        e=d.find_element_by_class_name("ci")
        time.sleep(1)
        e.click()
        time.sleep(3.4)
        w=len(d.find_elements_by_xpath('//div[@id="rs"]/table/tbody/tr'))
        q=len(d.find_elements_by_xpath('//div[@id="rs"]/table/tbody/tr/th'))
        time.sleep(0.5)
        g=0
        c=0
        try:
            if gkn=="RMR501712.xlsx":
                for m in range(2,w):
                    sums=0
                    for j in range(1,q):
                        #time.sleep(0.05)
                        try:
                            time.sleep(0.2)
                            u=d.find_element_by_xpath('//div[@id="rs"]/table/tbody/tr[{0}]/td[{1}]'.format(str(m),str(2))).text
                            a=(d.find_element_by_xpath('//div[@id="rs"]/table/tbody/tr[{0}]/td[{1}]'.format(str(m),str(j))).text)

                            #a=(d.find_element_by_xpath('//div[@id="rs"]/table/tbody/tr[{0}]/td[{1}]'.format(str(m),str(j))).text)
                            if j == 1 or j==7:
                                g+=1
                                if j==1:
                                    jka.append(a)
                                    #l.append(a)
                            elif j==2 and gkn=="RMR501712.xlsx":
                                l.append(i)
                                if u=="SOCIALLY RELEVANT PROJECT":
                                    continue
                                jka.append(a)
                                l.append(u)
                                bo.add(u)
                                c+=1
                                g+=1
                            elif j==3 and gkn=="RMR501712.xlsx":
                                l.append(a)
                                g+=1
                                jka.append(a)
                            elif j==4 and gkn=="RMR501712.xlsx":
                                l.append(a)
                                g+=1
                                jka.append(a)
                            elif j==5 and gkn=="RMR501712.xlsx":
                                l.append(a)
                                g+=1
                                jka.append(a)
                                es+=int(a)
                            elif j==6 and gkn=="RMR501712.xlsx":
                                l.append(a)
                                jka.append(a)
                                if a=="F":
                                    dert.append([u,i])
                                if zx[a]==0 and gkn=="RMR501712.xlsx":
                                    if tr[c]==0:
                                        tr[c]=1
                                    else:
                                        tr[c]=tr[c]+1
                                g+=1
                            elif j==8 and gkn=="RMR501712.xlsx":
                                l.append(a)
                                jka.append(a)
                                sums=sums+(jk[a]*fg[int(c)])
                                l.append(sums)
                                zf.append(int(sums))
                                g+=1
                            if (int(g%8==0) and gkn=="RMR501712.xlsx"):
                                if c==1 and gkn=="RMR501712.xlsx":
                                    q10.append(l)

                                    l.clear()
                                elif c==2 and gkn=="RMR501712.xlsx":
                                    q1.append(l)

                                    l.clear()
                                elif c==3 and gkn=="RMR501712.xlsx":
                                    q2.append(l)
                                    l.clear()
                                elif c==4 and gkn=="RMR501712.xlsx":
                                    q3.append(l)

                                    l.clear()
                                elif c==5 and gkn=="RMR501712.xlsx":
                                    q4.append(l)

                                    l.clear()
                                elif c==6 :
                                    q5.append(l)

                                    l.clear()
                                elif c==7 and gkn=="RMR501712.xlsx":
                                    q6.append(l)

                                    l.clear()
                                elif c==8:
                                    q7.append(l)

                                    l.clear()
                                elif c==9:
                                    q8.append(l)

                                    l.clear()
                                elif c==10:
                                    q9.append(l)

                                    l.clear()
                                elif c==11:
                                    q10.append(l)
                                    l.clear()
                                elif c==12:
                                    q11.append(l)
                                    l.clear()
                                elif c==13:
                                    q12.append(l)
                                    l.clear()
                                elif c==14:
                                    q13.append(l)
                                    l.clear()
                                elif c==15:
                                    q14.append(l)
                                    l.clear()
                                elif c==16:
                                    q15.append(l)
                                    l.clear()
                        except exceptions.StaleElementReferenceException:
                            time.sleep(0.1)
                            l.clear()
                            jka.clear()
                            time.sleep(0.5)
                            u=d.find_element_by_xpath('//div[@id="rs"]/table/tbody/tr[{0}]/td[{1}]'.format(str(m),str(2))).text
                            time.sleep(1)
                            a=(d.find_element_by_xpath('//div[@id="rs"]/table/tbody/tr[{0}]/td[{1}]'.format(str(m),str(j))).text)
                            if j == 1 or j==7:
                                g+=1
                                if j==1:
                                    jka.append(a)
                                    #l.append(a)
                            if j==2:
                                l.append(i)
                                jka.append(a)
                                if u=="SOCIALLY RELEVANT PROJECT":
                                    continue
                                l.append(u)
                                bo.add(u)
                                zvbm+=1
                                c+=1
                            if j==3:
                                l.append(a)
                                jka.append(a)
                                g+=1
                            if j==4:
                                l.append(a)
                                jka.append(a)
                                g+=1
                            if j==5:
                                l.append(a)
                                jka.append(a)
                                g+=1
                            if j==6:
                                l.append(a)
                                jka.append(a)
                                if a=="F":
                                    dert.append([u,i])
                                if zx[a]==0:
                                    if tr[c]==0:
                                        tr[c]=1
                                    else:
                                        tr[c]=tr[c]+1
                                g+=1
                            if j==8:
                                l.append(a)
                                jka.append(a)
                                sums=sums+(jk[a]*fg[c])
                                l.append(sums)
                                zf.append(int(sums))
                                g+=1
                            if int(g%8==0) :
                                if c==1:
                                    q10.append(l)
                                    l.clear()
                                if c==2:
                                    q1.append(l)
                                    l.clear()
                                if c==3:
                                    q2.append(l)
                                    l.clear()
                                if c==4:
                                    q3.append(l)
                                    l.clear()
                                elif c==5:
                                    q4.append(l)
                                    l.clear()
                                elif c==6:
                                    q5.append(l)
                                    l.clear()
                                elif c==7:
                                    q6.append(l)
                                    l.clear()
                                elif c==8:
                                    q7.append(l)
                                    l.clear()
                                elif c==9:
                                    q8.append(l)
                                    l.clear()
                                elif c==10:
                                    q9.append(l)
                                    l.clear()
                                elif c==11:
                                    q10.append(l)
                                    l.clear()
                                elif c==12:
                                    q11.append(l)
                                    l.clear()
                                elif c==13:
                                    q12.append(l)
                                    l.clear()
                                elif c==14:
                                    q13.append(l)
                                    l.clear()
                                elif c==15:
                                    q14.append(l)
                                    l.clear()
                                elif c==16:
                                    q15.append(l)
                                    l.clear()
                zvbm+=1
                fgkl=[]
                if zvbm==1 and gkn=="RMR501712.xlsx":
                        hjk.insert(0,"hallticketnumber")
                        for il in range(0,q):
                            hjk.append("subjectcode")
                            hjk.append("subjectname")
                            hjk.append("internal")
                            hjk.append("external")
                            hjk.append("total")
                            hjk.append("resultstatus")
                            hjk.append("grade")
                        hjk.append("total")
                        hjk.append("gpa")
                        ws.append(hjk)
                t.clear()
                ak.append([i,(sum(zf)/sum(ery))])
                lkj.append([(sum(zf)/sum(ery)),i])
                jka.append(es)
                jka.append((sum(zf)/sum(ery)*10))
                ws.append(jka)
            else:
                return render(request,"z.html")
        except selenium.common.exceptions.WebDriverException:
            continue
    print(lkj)
    lkj.sort()
    lkj.reverse()
    bx=1
    for i in lkj:
        i.insert(0,bx)
        q17.append(i)
        bx+=1
    for i in ak:
        q15.append(i)
    dert.sort()
    for i in dert:
        q16.append(i)
    qm=list(bo)
    yu=dict()
    for i in range(1,len(qm)):
        yu[qm[i-1]]=tr[i]
    for i,j in yu.items():
        q11.append([i,int((j/op)*10),int(op-j)])
    print(yu)
    v=Reference(q11,min_col=1,min_row=2,max_col=3,max_row=15)
    ch=BarChart()
    ch.title="  PASS PERCENTAGE OF VARIOUS SUBJECTS     "
    ch.x_axis.title="   SUBJECTS    "
    ch.y_axis.title="   PASS PERCENTAGE "
    ch.add_data(v)
    q11.add_chart(ch,"G2")
    rw=str(chr(92))
    dbs=sp.split(rw)
    au=""
    for i in range(0,len(dbs)):
        au=au+dbs[i]+rw+rw
    m="RMRD501712.xlsx"
    au=au+m
    if gkn=="RMR501712.xlsx":
        wb.save(au)
    adf=cvb.find(",")
    cm=[]
    if adf==-1:
        cm.append(cvb)
    else:
        cm=cvb.split(",")

    if gkn== "RMR501712.xlsx":
        for i in range(0,len(cm)):
            f=open(au,"rb")
            temp=render_to_string("d.html")
            email=EmailMessage("hi user from RMRD501712",temp,settings.EMAIL_HOST_USER,[cm[i]],)
            email.content_subtype="xlsx"
            email.attach(m,f.read())
            email.fail_silently=False
            email.send()
    #mn.clear()
    return render(request,"b.html")
def z(request):
    n=request.POST["n"]
    e=request.POST['e']
    f=request.POST['f']
    l=mg(name=n,email=e,feed=f)
    l.save()
    return redirect("/")